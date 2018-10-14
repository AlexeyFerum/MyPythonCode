from flask import Flask, request, session
from openpyxl import load_workbook
import os

app = Flask(__name__)
# нужен для session
# session - переменная в которой хранится информация о прошлых действиях пользователя
app.config['SECRET_KEY'] = "secretkey"


@app.route("/course", methods = ['GET', 'POST'])
def course():
    # Получение значения ссылки методом Get
    st = request.args.get('val', '')
    session['course'] = st
    return '''
        <!doctype html>
        <title>Excel</title>
        <h1>Выберите группу</h1>
        <ul>
        <li><a href="/group?val=Группа1">Группа 1</a></li>
        <li><a href="/group?val=Группа2">Группа 2</a></li>
        </ul>
        '''


@app.route("/group", methods=['GET', 'POST'])
def group():
    st = request.args.get('val', '')
    session['group'] = st
    dir = session['course']+'\\'+session['group']
    # Получение всех файлов в директории
    files = os.listdir(dir)
    ws = []
    # Определение xslx файлов и сразу их открытие с помощью библиотеки openpyxl
    for f in files:
        if f.find(".xlsx") != -1:
            workbook = load_workbook(dir + '\\'+f, data_only=True)
            ws.append(workbook.get_active_sheet())
            workbook.close()
    html_data = """
    <html>
        <head>
            <title>
            Результаты 
            </title>
        </head>
        <body>
        <h3>
        Таблица
        """
    html_data += session['course'] + ' ' + session['group']
    html_data += """
        </h3>
        <table>
    """
    if len(ws) != 0:
        # Заголовок таблицы
        for i in range(1, 3):
            html_data += "<tr>"
            for j in range(1, 3):
                if ws[0].cell(row=i, column=j).value is None:
                    html_data += "<td>" + ' ' + "</td>"
                else:
                    html_data += "<td>" + str(ws[0].cell(row=i, column=j).value) + "</td>"
            for w in range(0, len(ws)):
                for j in range(3, ws[w].max_column+1):
                    if ws[w].cell(row=i, column=j).value is not None:
                        html_data += "<td colspan='4'>" + str(ws[w].cell(row=i, column=j).value) + "</td>"
            html_data += "</tr>"

        for i in range(3, ws[0].max_row):
            html_data += "<tr>"
            # Первая таблица с двумя повторяющимися столбцами
            for j in range(1, ws[0].max_column+1):
                if ws[0].cell(row=i, column=j).value is None:
                    html_data += "<td>" + "</td>"
                else:
                    html_data += "<td>" + str(ws[0].cell(row=i, column=j).value) + "</td>"
            # Остальные таблицы
            for w in range(1, len(ws)):
                for j in range(3, ws[w].max_column+1):
                    if ws[w].cell(row=i, column=j).value is None:
                        html_data += "<td>" + "</td>"
                    else:
                        html_data += "<td>" + str(ws[w].cell(row=i, column=j).value) + "</td>"
            html_data += "</tr>"
    html_data += '''</table>
    </body>
    </html>
'''
    print(html_data)
    return html_data


# Начальная страница одиночный /
@app.route('/')
def first():
    return '''
    <!doctype html>
    <title>Excel</title>
    <h1>Выберите курс</h1>
    <ul>
    <li> <a href="/course?val=Курс1">Курс 1</a></li>
    <li> <a href="/course?val=Курс2">Курс 2</a></li>
    </ul>
    '''


if __name__ == "__main__":
    app.debug = True
    app.run()
    session.clear()
